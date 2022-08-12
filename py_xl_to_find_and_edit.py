from openpyxl import load_workbook

file_name='H00-ZA-E-86325_S1m.xlsx'

sheetname = 'HMB Table'

#load excel file
workbook = load_workbook(filename=file_name)

#sheet 1 changes
sheet1 = workbook[sheetname]

i = int(input("Enter any number except 0 to search : "))

while(i!=0):
    streamno = int(input("Enter stream no. : "))
    stream = "Stream no. " + str(streamno)
    for j in range(1, 1606):
        if(sheet1.cell(row=j, column=1).value==stream):
            #common for V and L
            print("mass flow rate : ",sheet1.cell(row=j+4, column=3).value)
            prev_fr = sheet1.cell(row=j+4, column=3).value
            ch = int(input("do u wanna continur : "))
            if(ch==1):
                now_fr = float(input("Enter new flowrate :" ))
                sheet1.cell(row=j+4, column=3).value = now_fr
                sheet1.cell(row=j+8, column=3).value = now_fr
                
                    #molar flow calc
                temp = sheet1.cell(row=j+4, column=3).value/sheet1.cell(row=j+6, column=3).value
                sheet1.cell(row=j+5, column=3).value = temp
                sheet1.cell(row=j+9, column=3).value = temp
                
                
                if(sheet1.cell(row=j+7, column=2).value == "Vapor phase"):        
                    # for gas
                    sheet1.cell(row=j+18, column=6).value = now_fr
                    sheet1.cell(row=j+18, column=11).value = temp

                    prev_nfr = sheet1.cell(row=j+10, column=3).value
                    temp3 = prev_nfr * now_fr/prev_fr
                    sheet1.cell(row=j+10, column=3).value = temp3

                    temp4 = now_fr / sheet1.cell(row=j+8, column=10).value
                    sheet1.cell(row=j+11, column=3).value = temp4

                else:
                    sheet1.cell(row=j+17, column=6).value = now_fr
                    sheet1.cell(row=j+17, column=11).value = temp

                    temp4 = now_fr / sheet1.cell(row=j+8, column=10).value
                    sheet1.cell(row=j+10, column=3).value = temp4

                    prev_nfr = sheet1.cell(row=j+11, column=3).value
                    temp3 = prev_nfr * now_fr/prev_fr
                    sheet1.cell(row=j+11, column=3).value = temp3


                #enthalpy
                temp2 = sheet1.cell(row=j+8, column=10).value * temp / (1e06)
                sheet1.cell(row=j+4, column=10).value = temp2
            else:
                continue
            print("--------------------------------------------")
            print("\n")

    i = int(input("Enter any number except 0 to search again : "))

workbook.save(filename=file_name)


