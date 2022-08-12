from openpyxl import load_workbook

file_name='H00-ZA-E-86348_S1 MASTER.xlsx'

sheetname = 'HMB Table'

#load excel file
#onlt use data only to ignore formula
workbook = load_workbook(filename=file_name, data_only=True)

#sheet 1 changes
sheet1 = workbook[sheetname]

i = int(input("Enter any number except 0 to search : "))

while(i!=0):
    streamno = int(input("Enter stream no. : "))
    stream = "Stream no. " + str(streamno)
    for j in range(1, 753):
        if(sheet1.cell(row=j, column=1).value==stream):
            #common for V and L
            print("mass flow rate : ",sheet1.cell(row=j+4, column=3).value)
            print("molar flow rate : ",sheet1.cell(row=j+5, column=3).internal_value)
               
            if(sheet1.cell(row=j+7, column=2).value == "Vapor phase"):        
                # for gas
                print("Normal volume flow : ",sheet1.cell(row=j+10, column=3).value)
                print("volume flow : ",sheet1.cell(row=j+11, column=3).value)

            else:
                print("Volume flow : ",sheet1.cell(row=j+10, column=3).value)
                print("Std Volume flow : ",sheet1.cell(row=j+11, column=3).value)

    print("--------------------------------------------")
    print("\n")

    i = int(input("Enter any number except 0 to search again : "))

#workbook.save(filename=file_name)


