from openpyxl import load_workbook

file_name='H10-SA-E-86908_02.xlsx'

#load excel file
workbook = load_workbook(filename=file_name, data_only=True)
sheetnames = workbook.sheetnames
total = len(sheetnames)
total = total - 3
print("Total sheets = ", total)
j=1

for i in sheetnames:
    # print(i,"\n")
    temp = workbook[i]
    # print(temp.cell(row=5, column=26).value,"\n")
    # temp.cell(row=5, column=26).value = j
    # temp.cell(row=5, column=28).value = total
    # j=j+1
    if(temp.cell(row=6, column=30).value == "02"):
        print("-----------   ",i,"   ------------------")

# workbook.save(filename=file_name)




# index = workbook["Index"]
# indexdata = ["No Sheet in index" for x in range(len(sheetnames))]
# j=0

# for i in range(9 , 57):
#     indexdata[j] = index.cell(row=i, column=2).value
#     j=j+1

# j=48
# for i in range(9 , 41):
#     indexdata[j] = index.cell(row=i, column=13).value
#     j=j+1

# sheetdata = ["" for x in range(len(sheetnames))]

# for i in range(len(sheetnames)):
#     temp = workbook[sheetnames[i]]
#     sheetdata[i]=temp.cell(row=8, column=9).value

# count=1
# # for i,j,z in zip(indexdata,sheetnames,sheetdata):
# #     print(count,". ",i,"  ----->  ",j, " ------->   ",z)
# #     count = count + 1
# print("-----------------------------------")
# for i in sheetnames:
#     print(i)
# print("-----------------------------------")
# print("-----------------------------------")
# for i in indexdata:
#     print(i)
# print("-----------------------------------")
# print("-----------------------------------")
# for i in sheetdata:
#     print(i)
# print("-----------------------------------")
# print("-----------------------------------")
